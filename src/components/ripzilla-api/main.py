import io
import os
import re
import colorsys

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from fastapi import FastAPI, UploadFile, File, HTTPException, Form
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse
from supabase import create_client, Client

app = FastAPI()

app.add_middleware(
    CORSMiddleware,
    allow_origins=["https://ripzillatcg.com", "http://localhost:5173"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# ── SUPABASE STORAGE ──────────────────────────────────────────────────────────
SUPABASE_URL = os.environ.get("SUPABASE_URL", "https://qzvhqslwhgzaztuypxhq.supabase.co")
SUPABASE_KEY = os.environ.get("SUPABASE_SERVICE_KEY")
BUCKET       = "card-sets"

supabase: Client = create_client(SUPABASE_URL, SUPABASE_KEY)

def storage_list() -> list:
    res = supabase.storage.from_(BUCKET).list()
    return [f["name"] for f in res if f["name"].endswith((".xlsx", ".xls"))]

def storage_download(filename: str) -> bytes:
    res = supabase.storage.from_(BUCKET).download(filename)
    return res

def storage_upload(filename: str, data: bytes):
    # upsert so re-uploading same name overwrites cleanly
    supabase.storage.from_(BUCKET).upload(
        filename, data,
        {"content-type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", "upsert": "true"}
    )

# ── SHARED HELPERS ────────────────────────────────────────────────────────────
TYPE_COLORS = {
    "Grass":      ("4CAF50", "000000"),
    "Fire":       ("FF5722", "FFFFFF"),
    "Water":      ("2196F3", "FFFFFF"),
    "Lightning":  ("FFC107", "000000"),
    "Psychic":    ("9C27B0", "FFFFFF"),
    "Fighting":   ("FF9800", "000000"),
    "Darkness":   ("424242", "FFFFFF"),
    "Metal":      ("9E9E9E", "000000"),
    "Dragon":     ("3F51B5", "FFFFFF"),
    "Colorless":  ("BDBDBD", "000000"),
    "Rainbow E":  ("E91E63", "FFFFFF"),
    "Psychic E":  ("9C27B0", "FFFFFF"),
    "Grass E":    ("4CAF50", "000000"),
    "Fighting E": ("FF9800", "000000"),
}
DEFAULT_TYPE   = ("607D8B", "FFFFFF")
ETB_CELL_COLOR = "ADD8E6"

NON_POKEMON_TYPES = {
    "I", "PT", "Su", "St", "Grass E", "Fire E", "Water E",
    "Lightning E", "Psychic E", "Fighting E", "Darkness E",
    "Metal E", "Dragon E", "Colorless E", "Rainbow E", "Double Colorless E",
}

RARITY_FILL = {
    "Common":                    "FFFFFF",
    "Uncommon":                  "E8F5E9",
    "Rare":                      "E3F2FD",
    "Double Rare":               "EDE7F6",
    "Illustration Rare":         "FFF8E1",
    "Ultra Rare":                "FFE0B2",
    "Special Illustration Rare": "FCE4EC",
    "Mega Attack Rare":          "E8EAF6",
    "Mega Hyper Rare":           "F3E5F5",
}

def make_border():
    t = Side(style="thin", color="CCCCCC")
    return Border(left=t, right=t, top=t, bottom=t)

def generate_username_colors(usernames):
    unique = sorted(set(usernames))
    n = len(unique)
    color_map = {}
    for i, username in enumerate(unique):
        hue = i / n if n else 0
        r, g, b = colorsys.hls_to_rgb(hue, 0.82, 0.65)
        color_map[username] = "{:02X}{:02X}{:02X}".format(int(r*255), int(g*255), int(b*255))
    return color_map

def text_color_for_bg(hex_bg):
    r, g, b = int(hex_bg[0:2],16), int(hex_bg[2:4],16), int(hex_bg[4:6],16)
    return "000000" if (0.299*r + 0.587*g + 0.114*b)/255 > 0.5 else "FFFFFF"

# ── CARD SET PARSER ───────────────────────────────────────────────────────────
def clean_card_name(name, card_type):
    if card_type in NON_POKEMON_TYPES:
        return name
    name = re.sub(r"^[\w\s]+'s\s+", "", name)
    name = re.sub(r"^Mega\s+", "", name)
    name = re.sub(r"\s+[XY]$", "", name)
    name = re.sub(r"ex$", "", name)
    return name.strip()

def parse_card_input(raw):
    rows = []
    for line in raw.splitlines():
        line = line.strip()
        if not line:
            continue
        parts = line.split("\t")
        if len(parts) < 5:
            continue
        no, mark, name, card_type, rarity = parts[:5]
        if no.lower() in ("no.", "no"):
            continue
        rows.append((no.strip(), mark.strip(), clean_card_name(name.strip(), card_type.strip()), card_type.strip(), rarity.strip()))
    return rows

def build_card_excel(rows):
    wb = Workbook()
    ws = wb.active
    ws.title = "Card List"
    border = make_border()
    for col, h in enumerate(["No.", "Mark", "Card Name", "Type", "Rarity"], 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font      = Font(bold=True, color="FFFFFF", name="Arial", size=11)
        cell.fill      = PatternFill("solid", start_color="1A237E")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = border
    ws.row_dimensions[1].height = 22
    for r, (no, mark, name, card_type, rarity) in enumerate(rows, 2):
        row_bg = RARITY_FILL.get(rarity, "FFFFFF")
        bg_color, fg_color = TYPE_COLORS.get(card_type, DEFAULT_TYPE)
        for col, val in enumerate([no, mark, name, card_type, rarity], 1):
            cell = ws.cell(row=r, column=col, value=val)
            cell.border    = border
            cell.alignment = Alignment(horizontal="left" if col == 3 else "center", vertical="center")
            if col == 4:
                cell.fill = PatternFill("solid", start_color=bg_color)
                cell.font = Font(name="Arial", size=10, color=fg_color)
            else:
                cell.fill = PatternFill("solid", start_color=row_bg)
                cell.font = Font(name="Arial", size=10)
        ws.row_dimensions[r].height = 18
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 8
    ws.column_dimensions["C"].width = 30
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 26
    ws.freeze_panes = "A2"
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf

# ── ORDER PROCESSING ──────────────────────────────────────────────────────────
def clean_product_name(name):
    name = re.sub(r"💎+", "", name).strip()
    etb_tag = " (ETB)" if re.search(r"\(.*?ETB.*?\)", name, flags=re.IGNORECASE) else ""
    name = re.sub(r"\(.*?ETB.*?\)", "", name, flags=re.IGNORECASE).strip()
    name = re.sub(r"^[\w\s]+'s\s+", "", name)
    name = re.sub(r"^Mega\s+", "", name)
    name = re.sub(r"\s+[XY]$", "", name)
    name = re.sub(r"\s*\bex\b\s*$", "", name, flags=re.IGNORECASE)
    return name.strip() + etb_tag

def is_etb(raw):
    return "ETB" in raw.upper()

def parse_orders(csv_bytes):
    df = pd.read_csv(io.BytesIO(csv_bytes))
    df = df[["Title", "Product", "Username"]].copy()
    df["_raw"]   = df["Product"]
    df["is_etb"] = df["_raw"].apply(is_etb)
    df["Product"] = df["_raw"].apply(clean_product_name)
    etb_users  = set(df.loc[df["is_etb"], "Username"])
    etb_df     = df[df["Username"].isin(etb_users)].drop(columns=["_raw"]).sort_values("Username").reset_index(drop=True)
    regular_df = df[~df["Username"].isin(etb_users)].drop(columns=["_raw"]).sort_values("Username").reset_index(drop=True)
    return regular_df, etb_df

def load_card_file_from_bytes(data: bytes, label: str) -> pd.DataFrame:
    df = pd.read_excel(io.BytesIO(data), dtype=str)
    df.columns = [c.strip() for c in df.columns]
    df["_name_key"] = df["Card Name"].str.strip().str.lower()
    return df[["No.", "_name_key", "Card Name", "Type", "Rarity"]]

def join_card_tables(file1_bytes, file1_name, file2_bytes, file2_name):
    label1 = os.path.splitext(file1_name)[0]
    label2 = os.path.splitext(file2_name)[0]
    df1 = load_card_file_from_bytes(file1_bytes, label1)
    df2 = load_card_file_from_bytes(file2_bytes, label2)
    merged = pd.merge(
        df1.rename(columns={"No.": f"{label1} No.", "Card Name": "_name1", "Type": "_type1", "Rarity": "_rarity1"}),
        df2.rename(columns={"No.": f"{label2} No.", "Card Name": "_name2", "Type": "_type2", "Rarity": "_rarity2"}),
        on="_name_key", how="outer",
    )
    merged["Card Name"] = merged["_name1"].combine_first(merged["_name2"])
    merged["Type"]      = merged["_type1"].combine_first(merged["_type2"])
    merged[f"{label1} No."] = merged[f"{label1} No."].fillna("—")
    merged[f"{label2} No."] = merged[f"{label2} No."].fillna("—")
    result = merged[["Card Name", f"{label1} No.", f"{label2} No.", "Type"]]
    result = result.sort_values("Card Name", key=lambda s: s.str.lower()).reset_index(drop=True)
    return result, [f"{label1} No.", f"{label2} No."]

def build_card_index(cards_df, set_cols):
    cards_df = cards_df.copy()
    cards_df["_key"] = cards_df["Card Name"].str.strip().str.lower()
    cards_agg = (
        cards_df.groupby("_key", sort=False)
        .agg(**{
            "Card Name": ("Card Name", "first"),
            "Type":      ("Type",      "first"),
            **{col: (col, lambda s: ", ".join(v for v in s if v != "—")) for col in set_cols},
        })
        .reset_index()
    )
    for col in set_cols:
        cards_agg[col] = cards_agg[col].replace("", "—")
    return cards_agg

def process_orders(orders, cards_agg, set_cols):
    orders = orders.copy()
    orders["_key"] = orders["Product"].str.strip().str.lower()
    orders["_key"] = orders["_key"].str.replace(r"\s*\(etb\)\s*$", "", regex=True)
    merged = orders.merge(cards_agg[["_key", "Type"] + set_cols], on="_key", how="left")
    for col in set_cols:
        merged[col] = merged[col].fillna("—")
    merged["Type"] = merged["Type"].fillna("")
    return merged

def write_report_sheet(ws, df, header_color, set_cols):
    border = make_border()
    user_colors  = generate_username_colors(df["Username"].tolist())
    display_cols = ["Title", "Username", "Product"] + set_cols + ["Type"]
    col_widths   = [55, 20, 28] + [max(len(c)+4, 16) for c in set_cols] + [16]
    for c, h in enumerate(display_cols, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font      = Font(bold=True, color="FFFFFF", name="Arial", size=11)
        cell.fill      = PatternFill("solid", start_color=header_color)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = border
    ws.row_dimensions[1].height = 22
    type_col_idx = display_cols.index("Type") + 1
    prod_col_idx = display_cols.index("Product") + 1
    for r, row in enumerate(df[display_cols].itertuples(index=False), 2):
        username   = row[display_cols.index("Username")]
        user_bg    = user_colors.get(username, "FFFFFF")
        user_fg    = text_color_for_bg(user_bg)
        product    = row[display_cols.index("Product")]
        is_etb_row = "(ETB)" in str(product)
        card_type  = row[display_cols.index("Type")]
        type_bg, type_fg = TYPE_COLORS.get(str(card_type), DEFAULT_TYPE)
        for c, val in enumerate(row, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.border    = border
            cell.alignment = Alignment(horizontal="left" if c in (prod_col_idx, type_col_idx) else "center", vertical="center")
            if c == type_col_idx:
                cell.fill = PatternFill("solid", start_color=type_bg)
                cell.font = Font(name="Arial", size=10, color=type_fg)
            elif c == prod_col_idx and is_etb_row:
                cell.fill = PatternFill("solid", start_color=ETB_CELL_COLOR)
                cell.font = Font(name="Arial", size=10, color="000000", bold=True)
            else:
                cell.fill = PatternFill("solid", start_color=user_bg)
                cell.font = Font(name="Arial", size=10, color=user_fg)
        ws.row_dimensions[r].height = 18
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w
    ws.freeze_panes = "A2"

# ── ROUTES ────────────────────────────────────────────────────────────────────
@app.get("/health")
def health():
    return {"status": "ok"}

@app.get("/sets")
def get_sets():
    try:
        files = storage_list()
        return {"sets": sorted(files)}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to list sets: {str(e)}")

@app.post("/sets/upload")
async def upload_set(file: UploadFile = File(...)):
    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="Please upload an Excel (.xlsx) file")
    contents = await file.read()
    try:
        storage_upload(file.filename, contents)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Upload failed: {str(e)}")
    return {"message": f"{file.filename} uploaded successfully", "filename": file.filename}

@app.post("/sets/parse")
async def parse_card_set(raw_text: str = Form(...), set_name: str = Form(...)):
    if not set_name.strip():
        raise HTTPException(status_code=400, detail="Set name is required")
    rows = parse_card_input(raw_text)
    if not rows:
        raise HTTPException(status_code=400, detail="No valid card data found. Make sure it is tab-separated.")
    buf = build_card_excel(rows)
    filename = f"{set_name.strip().replace(' ', '_')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

@app.post("/generate-report")
async def generate_report(
    file: UploadFile = File(...),
    set1: str = Form(...),
    set2: str = Form(...),
):
    if not file.filename.endswith(".csv"):
        raise HTTPException(status_code=400, detail="Please upload a CSV file")
    if set1 == set2:
        raise HTTPException(status_code=400, detail="Please select two different sets")

    try:
        set1_bytes = storage_download(set1)
        set2_bytes = storage_download(set2)
    except Exception as e:
        raise HTTPException(status_code=404, detail=f"Could not load sets from storage: {str(e)}")

    csv_bytes = await file.read()

    try:
        regular_df, etb_df = parse_orders(csv_bytes)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Error parsing CSV: {str(e)}")

    try:
        cards_df, set_cols = join_card_tables(set1_bytes, set1, set2_bytes, set2)
        cards_agg = build_card_index(cards_df, set_cols)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error joining card tables: {str(e)}")

    wb = Workbook()
    first = True
    for sheet_name, (orders_df, header_color) in {"All Orders": (regular_df, "1A237E"), "ETB Orders": (etb_df, "4A148C")}.items():
        joined = process_orders(orders_df, cards_agg, set_cols)
        if first:
            ws = wb.active; ws.title = sheet_name; first = False
        else:
            ws = wb.create_sheet(sheet_name)
        write_report_sheet(ws, joined, header_color, set_cols)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=break_report.xlsx"}
    )


# ── STREAM SCRAPING ───────────────────────────────────────────────────────────
import httpx
from bs4 import BeautifulSoup
from datetime import datetime, timezone
from pydantic import BaseModel
from typing import Optional

SUPABASE_ANON_KEY = os.environ.get("SUPABASE_ANON_KEY")

class StreamPayload(BaseModel):
    url: str
    # Manual fallback fields — only required if scraping fails
    title: Optional[str] = None
    starts_at: Optional[str] = None  # ISO string e.g. "2026-04-20T18:00:00"
    thumbnail: Optional[str] = None

def scrape_whatnot(url: str) -> dict:
    """
    Try to scrape title, start time, and thumbnail from a Whatnot stream URL.
    Returns a dict with keys: title, starts_at (ISO string), thumbnail.
    Raises ValueError if scraping fails or data can't be found.
    """
    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
        "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
    }
    response = httpx.get(url, headers=headers, follow_redirects=True, timeout=10)
    if response.status_code != 200:
        raise ValueError(f"Could not fetch page (status {response.status_code})")

    soup = BeautifulSoup(response.text, "html.parser")

    # Try Open Graph / meta tags first
    title = None
    starts_at = None
    thumbnail = None

    og_title = soup.find("meta", property="og:title") or soup.find("meta", attrs={"name": "title"})
    if og_title:
        title = og_title.get("content", "").strip()

    og_image = soup.find("meta", property="og:image")
    if og_image:
        thumbnail = og_image.get("content", "").strip()

    # Whatnot stores event time in JSON-LD or data attributes
    import json, re
    scripts = soup.find_all("script", type="application/ld+json")
    for script in scripts:
        try:
            data = json.loads(script.string or "")
            if isinstance(data, list):
                data = data[0]
            if data.get("startDate"):
                starts_at = data["startDate"]
            if data.get("name") and not title:
                title = data["name"]
            if data.get("image") and not thumbnail:
                img = data["image"]
                thumbnail = img if isinstance(img, str) else (img[0] if isinstance(img, list) else None)
        except Exception:
            continue

    # Fallback: look for time in page text patterns like "Apr 20 at 6:00 PM"
    if not starts_at:
        time_pattern = re.search(
            r'(\w{3}\s+\d{1,2}\s+at\s+\d{1,2}:\d{2}\s*[AP]M)',
            response.text
        )
        if time_pattern:
            try:
                parsed = datetime.strptime(time_pattern.group(1), "%b %d at %I:%M %p")
                parsed = parsed.replace(year=datetime.now().year, tzinfo=timezone.utc)
                starts_at = parsed.isoformat()
            except Exception:
                pass

    if not title:
        raise ValueError("Could not extract stream title from page")

    return {"title": title, "starts_at": starts_at, "thumbnail": thumbnail}


@app.post("/streams")
async def add_stream(payload: StreamPayload):
    """
    Add a stream to Supabase.
    Tries to scrape Whatnot first. If scraping fails and manual fields
    are provided, uses those instead. If neither works, returns 422 with
    scrape_failed=True so the frontend can show the manual form.
    """
    title      = payload.title
    starts_at  = payload.starts_at
    thumbnail  = payload.thumbnail
    scrape_failed = False

    # Try scraping
    try:
        scraped = scrape_whatnot(payload.url)
        title     = title     or scraped["title"]
        starts_at = starts_at or scraped["starts_at"]
        thumbnail = thumbnail or scraped["thumbnail"]
    except Exception as e:
        scrape_failed = True
        # If manual fields weren't provided either, tell frontend to show manual form
        if not title or not starts_at:
            return JSONResponse(
                status_code=422,
                content={
                    "scrape_failed": True,
                    "detail": f"Could not auto-detect stream info: {str(e)}. Please enter details manually."
                }
            )

    # Save to Supabase
    try:
        res = supabase.table("streams").insert({
            "url":        payload.url,
            "title":      title,
            "starts_at":  starts_at,
            "thumbnail":  thumbnail,
        }).execute()
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to save stream: {str(e)}")

    return {"success": True, "scrape_failed": scrape_failed, "stream": res.data[0] if res.data else {}}


@app.get("/streams")
async def get_streams():
    """Return all upcoming streams ordered by start time."""
    try:
        now = datetime.now(timezone.utc).isoformat()
        res = supabase.table("streams").select("*").gte("starts_at", now).order("starts_at").execute()
        return {"streams": res.data}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to fetch streams: {str(e)}")


@app.delete("/streams/{stream_id}")
async def delete_stream(stream_id: str):
    """Delete a stream by ID."""
    try:
        supabase.table("streams").delete().eq("id", stream_id).execute()
        return {"success": True}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to delete stream: {str(e)}")


# ── STREAMS ───────────────────────────────────────────────────────────────────
import httpx
from bs4 import BeautifulSoup
from datetime import datetime, timezone
from pydantic import BaseModel
from typing import Optional

class StreamPayload(BaseModel):
    url: str
    title: Optional[str] = None
    starts_at: Optional[str] = None  # ISO string, provided if scraping fails

async def scrape_whatnot(url: str) -> dict:
    """
    Try to scrape title and start time from a Whatnot stream page.
    Returns dict with title and starts_at (ISO string) or raises on failure.
    """
    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
        "Accept-Language": "en-US,en;q=0.9",
    }
    async with httpx.AsyncClient(follow_redirects=True, timeout=10) as client:
        res = await client.get(url, headers=headers)
        res.raise_for_status()

    soup = BeautifulSoup(res.text, "html.parser")

    # Try og:title first, fall back to <title>
    title = None
    og_title = soup.find("meta", property="og:title")
    if og_title and og_title.get("content"):
        title = og_title["content"].strip()
    elif soup.title:
        title = soup.title.string.strip()

    # Try to find start time in JSON-LD or meta tags
    starts_at = None

    # Look for JSON-LD structured data
    for script in soup.find_all("script", type="application/ld+json"):
        try:
            import json
            data = json.loads(script.string)
            if isinstance(data, dict):
                for key in ("startDate", "startTime", "scheduledStartTime"):
                    if data.get(key):
                        starts_at = data[key]
                        break
        except Exception:
            continue

    # Try meta tags
    if not starts_at:
        for name in ("start_time", "event:start_time", "scheduledStartTime"):
            tag = soup.find("meta", attrs={"name": name}) or soup.find("meta", property=name)
            if tag and tag.get("content"):
                starts_at = tag["content"]
                break

    return {"title": title, "starts_at": starts_at}


@app.post("/streams/scrape")
async def scrape_stream(payload: StreamPayload):
    """
    Try to scrape a Whatnot URL for title and start time.
    Returns what was found — frontend uses this to pre-fill the form.
    Falls back gracefully if scraping fails.
    """
    result = {"title": None, "starts_at": None, "scraped": False, "error": None}
    try:
        data = await scrape_whatnot(payload.url)
        result["title"]     = data.get("title")
        result["starts_at"] = data.get("starts_at")
        result["scraped"]   = True
    except Exception as e:
        result["error"] = f"Could not scrape page: {str(e)}. Please enter details manually."
    return result


@app.post("/streams")
async def save_stream(payload: StreamPayload):
    """
    Save a stream to Supabase. title and starts_at must be provided
    (either from scraping or manual entry).
    """
    if not payload.title:
        raise HTTPException(status_code=400, detail="Stream title is required")
    if not payload.starts_at:
        raise HTTPException(status_code=400, detail="Stream start time is required")
    if not payload.url:
        raise HTTPException(status_code=400, detail="Stream URL is required")

    try:
        # Parse and normalize to UTC ISO string
        dt = datetime.fromisoformat(payload.starts_at.replace("Z", "+00:00"))
        starts_at_iso = dt.isoformat()
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid start time format")

    try:
        supabase.table("streams").insert({
            "title":     payload.title,
            "url":       payload.url,
            "starts_at": starts_at_iso,
        }).execute()
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to save stream: {str(e)}")

    return {"message": "Stream saved successfully"}


@app.delete("/streams/{stream_id}")
async def delete_stream(stream_id: str):
    """Delete a stream by ID."""
    try:
        supabase.table("streams").delete().eq("id", stream_id).execute()
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to delete stream: {str(e)}")
    return {"message": "Stream deleted"}
